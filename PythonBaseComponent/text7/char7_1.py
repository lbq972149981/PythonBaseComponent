from os.path import isfile as isfile
from time import time as time
from os.path import isdir, join, splitext
from os import listdir, remove
from xlwt import *
from docx import Document
import tkinter.ttk
import tkinter
import xlrd
import pickle
import struct
import zlib
# import rarfile
import binascii
import openpyxl
import random
import hashlib
import zipfile
# import rarfile
import _md5
import os
# import openpyxl
import win32com
from win32com import client
import difflib


# eg. 7-1向文本文件中写入内容
f = open('D:\\Python\\abc\sample.txt', 'a+')
s = '文本文件的读取方法\n文本文件的写入方法\n'
f.write(s)
f.close()

s = '文本文件的读取方法\n文本文件的写入方法\n'
with open('D:\\Python\\abc\sample.txt', 'a+') as f:
    f.write(s)

# eg. 7-2读取并显示文本文件的前5个字节
fp = open('D:\\Python\\abc\sample1.txt', 'r')  # python 3.x
print(fp.read(5))
print(fp.read(7))
print(fp.seek(0))
print(fp.read(8), '\n')

# eg. 7-3读取并显示文本文件所有行
f = open('D:\\Python\\abc\sample2.txt', 'r')
while True:
    line = f.readline()
    if line == '':
        break
    print(line, end=' ')
f.close()
print('\n')
f = open('D:\\Python\\abc\sample2.txt', 'r')
li = f.readlines()
for line in li:
    print(line, end=' ')
f.close()
print('\n')

# eg. 7-4移动文件指针
s = '中国山东烟台SDIBT'
fp = open(r'D:\Python\\abc\sample1.txt', 'w')
fp.write(s)
fp.close()
fp = open(r'D:\Python\\abc\sample1.txt', 'r')
print(fp.read(3))
print(fp.seek(13))
print(fp.read(1))
print(fp.seek(15))
print(fp.read(1))
print(fp.seek(3))
print('\n')
# print(fp.read(1))1

# eg. 7-5 读取文本文件data.txt中所有整数，将其按升序排序后再写入文本文件data_asc.txt中
with open('D:\\Python\\abc\data.txt', 'r') as fp:
    data = fp.readlines()
data = [int(line.strip()) for line in data]
data.sort()
data = [str(i)+'\n' for i in data]
with open('D:\\Python\\abc\data_asc.txt', 'w') as fp:
    fp.writelines(data)

# eg. 7-6 编写程序保存为嘚demo6.py运行后生成文件demo6_new.py
# 其中内容与demo6.py一致，但每行的行尾加上了行号
filename = 'D:\Python\\abc\demo6.py'
with open(filename, 'r') as fp:
    lines = fp.readlines()
lines = [line.rstrip()+' '*(100-len(line))+'# '+str(index)+'\n' for index, line in enumerate(lines)]
with open(filename[:-3]+'_new.py', 'w') as fp:
    fp.writelines(lines)

# eg. 7-7Pytho程序中代码复用度检测
Result = {}
AllLines = []
FileName = r'FindIdentifiersFromPyFile.py'
# FileName = input('Please input the file to check, including full path:')

# Read the content if given file
# Remove blank lines
# Remove all the whitespace string of every line
# preserving only one space character between words or operators
# note: The last line does not contain the '\n' character


def PreOperate():
    global AllLines
    with open(FileName, 'r') as fp:
        for line in fp:
            line = ''.join(line.split())
            if line != '':
                AllLines.append(line)


# check if the current position is still the duplicated one
def IfHasDuplicated(Index1):
    for item in Result.values():
        for it in item:
            if Index1 == it[0]:
                return it[1]
    return False


# if the current line Index2 is in a span of duplicated lines, return True, else False
def IsInSpan(Index2):
    for item in Result.values():
        for i in item:
            if i[0] <= Index2 < i[0]+i[1]:
                return True
    return False


def MainCheck():
    global Result
    TotalLen = len(AllLines)
    Index1 = 0
    while Index1 < TotalLen-1:
        # speed up
        span = IfHasDuplicated(Index1)
        if span:
            Index1 += span
            continue
        Index2 = Index1+1
        while Index2 < TotalLen:
            # speed up, skip the duplicated lines
            if IsInSpan(Index2):
                Index2 += 1
                continue
            src = ''
            des = ''
            for i in range(10):
                if Index2+i >= TotalLen:
                    break
                src += AllLines[Index1+i]
                des += AllLines[Index2+i]
                if src == des:
                    t = Result.get(Index1, [])
                    for tt in t:
                        if tt[0] == Index2:
                            tt[1] = i+1
                            break
                    else:
                        t.append([Index2, i+1])
                    Result[Index1] = t
                else:
                    break
            t = Result.get(Index1, [])
            for tt in t:
                if tt[0] == Index2:
                    Index2 += tt[1]
                    break
            else:
                Index2 += 1
        # optimize the Result dictionart, remove the item with span<3
        Result[Index1] = Result.get(Index1, [])
        for n in Result[Index1][::-1]:  # Note: here must use the reverse slice
            if n[1] < 3:
                Result[Index1].remove(n)
        if not Result[Index1]:
            del Result[Index1]

        # Comoute the min span of duplicated codes of line Index1, modify the step
        # Index1
        a = [ttt[1] for ttt in Result.get(Index1, [[Index1, 1]])]
        if a:
            Index1 += max(a)
        else:
            Index1 += 1


# Output the result
def Output():
    print('-'*20)
    print('-'*20)
    print('Result:')
    for key, value in Result.items():
        print("The original line is :\n{0}".format(AllLines[key]))
        print("Its line number is {0}".format(key))
        print('The duplicated line numbers are:')
        for i in value:
            print('    Start:', i[0], '    Span:', i[1])
        print('-'*20)
    print('-'*20)


if isfile(FileName):
    start = time()
    PreOperate()
    MainCheck()
    Output()
    print('Time uesd:', time()-start, '\n')

# eg. 7-8使用pickle模块写入二进制文件
f = open('D:\\Python\\abc\sample_pickle.dat', 'wb')
n = 7
i = 13000000
a = 99.056
s = '中国人民 123abc'
lst = [[1, 2, 3], [4, 5, 6], [7, 8, 9]]
tu = (-5, 10, 8)
coll = {4, 5, 6}
dic = {'a': 'apple', 'b': 'banana', 'g': 'grape', 'o': 'orange'}
try:
    pickle.dump(n, f)
    pickle.dump(i, f)
    pickle.dump(a, f)
    pickle.dump(s, f)
    pickle.dump(lst, f)
    pickle.dump(tu, f)
    pickle.dump(coll, f)
    pickle.dump(dic, f)
except:
    print('写文件异常')
finally:
    f.close()


# eg. 7-9读取例7-8中写入二进制文件的内容
f = open('D:\\Python\\abc\sample_pickle.dat', 'rb')
n = pickle.load(f)
# print(n)
i = 0
while i < n:
    x = pickle.load(f)
    print(x)
    i = i+1
f.close()
print('\n')

# eg. 7-10使用struct模块写入二进制文件
n = 1300000000
x = 96.45
b = True
s = 'al@ 中国'
sn = struct.pack('if? ', n, x, b)
f = open('D:\\Python\\abc\sample_struct.dat', 'wb')
f.write(sn)
f.write(s.encode())
f.close()

# eg.7-11使用struct模块读取例7-10写入二进制文件的内容
f = open('D:\\Python\\abc\sample_struct.dat', 'rb')
sn = f.read(9)
tu = struct.unpack('if? ', sn)
print(tu)
n, x, b1 = tu
print('n=', n, 'x=', x, 'b1=', b1)
s = f.read(9)
# s = s.decode()
print('s=', s, '\n')

# eg. 7-12计算CRC32值
print(zlib.crc32('1234'.encode()))
print(zlib.crc32('111'.encode()))
print(zlib.crc32('SDIBT'.encode()))
print(binascii.crc32('SDIBT'.encode()), '\n')

# eg. 7-13计算文本文档中最长行的长度
f = open('D:\Python\\abc\mydata.txt', 'r')
allLineLens = [len(line.strip()) for line in f]
f.close()
longest = max(allLineLens)
print(longest)

f = open('D:\Python\\abc\mydata.txt', 'r')
longest = max(len(line.strip()) for line in f)
f.close()
print(longest, '\n')

# eg. 7-14计算字符串MD5值
md5value = hashlib.md5()
md5value.update('12345'.encode())
md5value = md5value.hexdigest()
print(md5value)
md5value = _md5.md5()
md5value.update('12345'.encode())
md5value = md5value.hexdigest()
print(md5value, '\n')


# eg. 7-15判断一个文件是否为GIF图像文件
def is_gif(fname):
    f1 = open(fname, 'rb')  # 书上是‘r’???
    first4 = tuple(f1.read(4))
    f1.close()
    return first4 == ('G', 'I', 'F', '8')


print(is_gif('D:\Python\\abc\picture1.jpg'), '\n')

# eg. 7-16比较两个文本文件是否相同
# A = open(r'D:\Python\\abc\mydata.txt')
# B = open(r'D:\Python\\abc\sample.txt')
# contextA = A.read()
# contextB = B.read()
# s = difflib.SequenceMatcher(lambda x: x==" ", contextA, contextB)
# result = s.get_opcodes()
# for tag, i1, i2, j1, j2 in result:
#     print("%s contextA[%d:%d]=%s contextB[%d:%d]=%s" %\
#           (tag, i1, i2, contextA[i1:i2], j1, j2, contextB[j1, j2]))


# eg. 7-17使用xlwt模块写入Excel文件
book = Workbook()
sheet1 = book.add_sheet("First")
al = Alignment()
al.horz = Alignment.HORZ_CENTER
al.vert = Alignment.VERT_CENTER
borders = Borders()
borders.bottom = Borders.THICK
style = XFStyle()
style.alignment = al
style.borders = borders
row0 = sheet1.row(0)
row0.write(0, 'test', style=style)
book.save(r'D:\\Python\\abc\test.xls')

# eg. 7-18使用xlrd模块读取Excel文件
book = xlrd.open_workbook(r'D:\Python\abc\test.xls')
sheet1 = book.sheet_by_name('First')
row0 = sheet1.row(0)
print(row0[0])
print(row0[0].value, '\n')

# eg. 7-19使用Pywin32操作Excel文件
# xlApp = win32com.client.Dispatch('Excel.Application')
# xlBook = xlApp.Workbooks.Open(r'D:\Python\\abc\\test.xls')
# xlSht = xlBook.Worksheets('sheet1')
# aaa = xlSht.Cells(1, 2).Value
# xlSht.Cells(2, 3).Value = aaa
# xlBook.Close(SaveChanges=1)
# del xlApp

# eg. 7-20检查Word文档的连续重复字
# filename = r'D:\Python\codes\python3\Test\Threading.doc'
# word = client.Dispatch('Word.Application')
# doc = word.Documents.Open(filename)
# content = str(doc.Content)
# doc.Close()
# word.Quit()
#
# repeatedWords = []
#
# lens = len(content)
# for i in range(lens-2):
#     ch = content[i]
#     ch1 = content[i+1]
#     ch2 = content[i+2]
#     if u'\u4e00' <= ch <= u'\u9fa5' or ch in (',', '.', '、'):
#         if ch == ch1 and ch+ch1 not in repeatedWords:
#             print(ch + ch1)
#             repeatedWords.append(ch+ch1)
#         elif ch == ch2 and ch+ch1+ch2 not in repeatedWords:
#             print(ch+ch1+ch2)
#             repeatedWords.append(ch+ch1+ch2)
# print('\n')

# eg. 7-21编写程序，进行文件夹增量备份
# import CopyDir

# eg.7-22统计质问文件夹大小以及文件和子文件夹数量
totalSize = 0
fileNum = 0
dirNum = 0


def visitDir(path):
    global totalSize
    global fileNum
    global dirNum
    for lists in os.listdir(path):
        sub_path = os.path.join(path, lists)
        if os.path.isfile(sub_path):
            fileNum = fileNum+1
            totalSize = totalSize+os.path.getsize(sub_path)
        elif os.path.isdir(sub_path):
            dirNum = dirNum+1
            visitDir(sub_path)


def main(path):
    if not os.path.isdir(path):
        print('Error："', path, '" is not a directory or does not exist.' )
        return
    visitDir(path)


def sizeConvert(size):
    K, M, G = 1024, 1024**2, 1024**3
    if size >= G:
        return str(size/G)+'G Bytes'
    elif size >= M:
        return str(size/M)+'M Bytes'
    elif size >= K:
        return str(size/K)+'K Bytes'
    else:
        return str(size)+'Bytes'


def output(path):
    print('The total size of '+path+' is:'+sizeConvert(totalSize)+'('+str(totalSize)+' Bytes)')
    print('The total number of files in '+path+' is:', fileNum)
    print('The total number of directories in '+path+' is:', dirNum, '\n')


if __name__ == '__main__':
    path = r'D:\Python\abc'
    main(path)
    output(path)

# eg. 7-23统计指定目录所有C++源程序文件中不重复代码行数
AllLines = []
NotRepeatedLines = []
file_num = 0
code_num = 0


def LinesCount(directory):
    global AllLines
    global NotRepeatedLines
    global file_num
    global code_num
    for filename in listdir(directory):
        temp = join(directory, filename)
        if isdir(temp):
            LinesCount(temp)
        elif temp.endswith('.cpp'):
            file_num += 1
            with open(temp, 'r') as fp:
                line = fp.readline()
                if not line:
                    break
                if line not in NotRepeatedLines:
                    NotRepeatedLines.append(line)
                code_num += 1
    return code_num, len(NotRepeatedLines)


path = r'D:\Ccodes'
print("代码总行数:{0[0]}, 非重复代码行数:{0[1]}".format(LinesCount(path)))
print("文件数量:{0}".format(file_num), '\n')

# eg.7-24递归删除指定文件夹中指定类型文件
filetypes = ['.tmp', '.log', '.obj', '.txt']


def delCertainFiles(directory):
    if not isdir(directory):
        return
    for filename in listdir(directory):
        temp = join(directory, filename)
        if isdir(temp):
            delCertainFiles(temp)
        elif splitext(temp)[1] in filetypes:
            remove(temp)
            print(temp, 'deleted.…')


def main():
    directory = r'D:\Python\abc\xyz'
    # directory = sys.argv[1]
    delCertainFiles(directory)


main()

# eg. 7-25使用扩展库openpyxl读写Excel 2007及更高版本Excel文件
# fn = r'D:\Python\abc\test1.xlsx'
# wb = openpyxl.workbook.Workbook()
# ws = wb.create_sheet(title='Hello World!')
# ws['A1'] = '这是第一个单元格'
# ws['B1'] = 3.1415926
# wb.save(fn)
# wb = openpyxl.load_workbook(fn)
# ws = wb.worksheets[1]
# print(ws['A1'].value)
# ws.append([1, 2, 3, 4, 5])
# ws.merge_cells('F2:F3')
# ws['F2'] = ' = sum(A2:E2)'
# for r in range(10, 15):
#     for c in range(3, 8):
#         _ = ws.merge_cells(row=r, column=c, value=r*c)
# wb.save(fn)


# def generateRandomInformation(filename):
#     workbook = openpyxl.Workbook()
#     worksheet = workbook.worksheets[0]
#     worksheet.append(['姓名', '课程', '成绩'])
#     first = tuple('赵钱孙李')
#     middle = tuple('伟昀琛东')
#     last = tuple('坤艳志')
#     subjects = ('语文', '数学', '英语')
#     for i in range(200):
#         line = []
#         r = random.randint(1, 100)
#         name = random.choice(first)
#         if r > 50:
#             name = name + random.choice(middle)
#         name = name + random.choice(last)
#         line.append(name)
#         line.append(random.choice(subjects))
#         line.append(random.randint(0, 100))
#         worksheet.append(line)
#     workbook.save(filename)
#
#
# def getResult(oldfile, newfile):
#     result = dict()
#     workbook = openpyxl.load_workbook(oldfile)
#     worksheet = workbook.worksheets[0]
#     for row in worksheet.rows[1:]:
#         name, subject, grade = row[0].value, row[1].value, row[2].value
#         t = result.get(name, [])
#         f = t.get(subject, ())
#         if grade > f:
#             t[subject] = grade
#             result[name] = t
#     workbook1 = openpyxl.Workbook()
#     worksheet1 = workbook1.worksheets[0]
#     worksheet1.append(['姓名', '课程', '成绩'])
#     for name, t in result.items():
#         for subject, grade in t.items():
#             worksheet1.append([name, subject, grade])
#     workbook1.save(newfile)
#
#
# if __name__ == '__main__':
#     oldfile = r'D:\Python\abc\test1.xlsx'
#     newfile = r'D:\Python\abc\result.xlsx'
#     generateRandomInformation(oldfile)
#     getResult(oldfile, newfile)

# eg. 7-26查看指定xip和rar压缩文件中的文件列表
# fp = zipfile.ZipFile(r'D:\Download')
# for g in fp.namelist():
#     print(f)
# fp.close()

# r = rarfile.RarFile(r'asp网站.rar')
# for f in r.namelist():
#     print(f)
# r.close()

# eg. 7-27小学口算题库生成器
columnsNumber = 4


def main(rowsNumber=20, grade=4):
    if grade < 3:
        operators = '+-'
        biggest = 20
    elif grade <= 4:
        operators = '+-*/'
        biggest = 100
    elif grade == 5:
        operators = '+-*/('
        biggest = 100

    document = Document()
    table = document.add_table(rows=rowsNumber, cols=columnsNumber)
    for row in range(rowsNumber):
        for col in range(columnsNumber):
            first = random.randint(1, biggest)
            second = random.randint(1, biggest)
            operator = random.choice(operators)
            if operator != '(':
                if operator == '-':
                    if first < second:
                        first, second = second, first
                r = str(first).ljust(2, ' ')+' '+operator+str(second).ljust(2, ' ')+'='
            else:
                third = random.randint(1, 100)
                while True:
                    o1 = random.choice(operators)
                    o2 = random.choice(operators)
                    if o1 != ')' and o2 != '(':
                        break
                rr = random.randint(1, 100)
                if rr > 50:
                    if o2 == '-':
                        if second < third:
                            second, third = third, second
                    r = str(second).ljust(2, ' ')+o1+'('+str(second).ljust(2, ' ')+o2+str(third).ljust(2, ' ')+')='
                else:
                    if o1 == '-':
                        if first < second:
                            first, second = second, first
                    r = '('+str(first).ljust(2, ' ')+o1+str(second).ljust(2, ' ')+')'+o2+str(third).ljust(2, ' ')+'='
            cell = table.cell(row, col)
            cell.text = r
    document.save('kousuan.docx')
    os.startfile('kousuan.docx')


if __name__ == '__main__':
    app = tkinter.Tk()
    app.title('KouSuan-----by GLARE')
    app['width'] = 300
    app['height'] = 150
    labelNumber = tkinter.Label(app, text='Number:', justify=tkinter.RIGHT, width=50)
    labelNumber.place(x=10, y=40, width=50, height=20)
    comboNumber = tkinter.ttk.Combobox(app, values=(100, 200, 300, 400, 500), width=50)
    comboNumber.place(x=70, y=40, width=50, height=20)
    labelGrade = tkinter.Label(app, text='Grade:', justify=tkinter.RIGHT, width=50)
    labelGrade.place(x=130, y=40, width=50, height=20)
    comboGrade = tkinter.ttk.Combobox(app, values=(1, 2, 3, 4, 5), width=50, height=20)
    comboGrade.place(x=200, y=40, width=50, height=20)


    def generate():
        number = int(comboNumber.get())
        grade = int(comboGrade.get())
        main(number, grade)
    buttonGenerate = tkinter.Button(app, text='GO', width=40, command=generate)
    buttonGenerate.place(x=130, y=90, width=40, height=30)
    app.mainloop()

